import openpyxl

def getRowCount(file, sheetName):
    """Return total number of rows in a sheet"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    """Return total number of columns in a sheet"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, colnum):
    """Read data from a given cell"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=colnum).value


def writeData(file, sheetName, rownum, colnum, data):
    """Write data to a given cell"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)






